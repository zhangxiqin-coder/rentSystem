"""
个人资产管理API
支持两种上报方式：
1. 余额上报：平台名 + 当前余额 + 累计收益 → 自动算出转入/转出净额
2. 转入/转出：平台名 + 金额 → 自动更新当前余额

收益按年份管理：每年首次上报时自动归档上年收益，页面显示当年收益。
"""
import json
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from decimal import Decimal
from datetime import datetime, timedelta
from typing import Optional

from app.database import get_db
from app.models import AssetPlatform, AssetRecord, User, AssetItem, AssetSnapshot, FixedAsset
from app.schemas import (
    AssetPlatformCreate, AssetPlatformUpdate, AssetPlatformResponse,
    AssetPlatformDetailResponse, AssetRecordResponse, AssetRecordCreate,
    AssetRecordUpdate,
    AssetSummaryResponse,
    AssetTrendResponse, AssetTrendPoint, PlatformTrendPoint,
    ZhaopingfeiYearSummary, ZhaopingfeiSummaryResponse,
    AssetItemCreate, AssetItemUpdate, AssetItemResponse,
    PortfolioSummaryResponse,
    AssetSnapshotCreate, AssetSnapshotResponse,
    FixedAssetCreate, FixedAssetUpdate, FixedAssetResponse
)
from app.core.deps import get_current_user

router = APIRouter()


def _get_platform_or_404(db: Session, platform_id: int, user_id: int) -> AssetPlatform:
    """获取平台并校验归属"""
    platform = db.query(AssetPlatform).filter(
        AssetPlatform.id == platform_id,
        AssetPlatform.owner_id == user_id
    ).first()
    if not platform:
        raise HTTPException(status_code=404, detail="平台不存在")
    return platform


def _parse_yearly_earnings(platform: AssetPlatform) -> dict:
    """解析yearly_earnings JSON字段"""
    if not platform.yearly_earnings:
        return {}
    try:
        return json.loads(platform.yearly_earnings)
    except (json.JSONDecodeError, TypeError):
        return {}


def _save_yearly_earnings(platform: AssetPlatform, data: dict):
    """保存yearly_earnings JSON字段"""
    platform.yearly_earnings = json.dumps(data, ensure_ascii=False)


def _check_year_rollover(platform: AssetPlatform, db: Session):
    """检查是否需要跨年归档，如果是则归档上年收益"""
    current_year = datetime.now().year
    if current_year > (platform.current_year or 2026):
        # 跨年了！归档当前收益
        yearly = _parse_yearly_earnings(platform)
        year_key = str(platform.current_year)
        yearly[year_key] = str(platform.total_earnings or Decimal('0'))
        _save_yearly_earnings(platform, yearly)
        
        # 重置当前年份收益
        platform.total_earnings = Decimal('0')
        platform.current_year = current_year
        db.flush()


# ==================== 平台管理 ====================

@router.get("/assets/platforms", response_model=list[AssetPlatformResponse])
async def list_platforms(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取当前用户的所有资产平台"""
    return db.query(AssetPlatform).filter(
        AssetPlatform.owner_id == current_user.id,
        AssetPlatform.is_active == True
    ).order_by(AssetPlatform.sort_order, AssetPlatform.id).all()


@router.post("/assets/platforms", response_model=AssetPlatformResponse)
async def create_platform(
    data: AssetPlatformCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """创建资产平台"""
    current_year = datetime.now().year
    platform = AssetPlatform(
        name=data.name,
        current_balance=data.current_balance,
        total_earnings=Decimal('0'),
        current_year=current_year,
        yearly_earnings='{}',
        sort_order=data.sort_order,
        owner_id=current_user.id
    )
    db.add(platform)
    db.commit()
    db.refresh(platform)
    return platform


@router.put("/assets/platforms/{platform_id}", response_model=AssetPlatformResponse)
async def update_platform(
    platform_id: int,
    data: AssetPlatformUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """更新资产平台"""
    platform = _get_platform_or_404(db, platform_id, current_user.id)
    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(platform, key, value)
    platform.updated_at = datetime.now()
    db.commit()
    db.refresh(platform)
    return platform


@router.delete("/assets/platforms/{platform_id}")
async def delete_platform(
    platform_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """删除资产平台（同时删除所有关联记录）"""
    platform = _get_platform_or_404(db, platform_id, current_user.id)
    db.query(AssetRecord).filter(AssetRecord.platform_id == platform_id).delete()
    db.delete(platform)
    db.commit()
    return {"message": "平台已删除"}


# ==================== 资产变动记录 ====================

@router.post("/assets/records", response_model=AssetRecordResponse)
async def create_asset_record(
    data: AssetRecordCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """创建资产变动记录（包含余额上报和转入/转出两种模式）"""
    platform = _get_platform_or_404(db, data.platform_id, current_user.id)

    # 检查是否需要跨年归档
    _check_year_rollover(platform, db)

    record = AssetRecord(
        platform_id=data.platform_id,
        owner_id=current_user.id,
        record_type=data.record_type,
        notes=data.notes,
        balance_before=platform.current_balance,
        earnings_before=platform.total_earnings
    )

    if data.record_type == "balance":
        # 余额+收益上报模式（同时填余额和累计收益）
        if data.reported_balance is None or data.reported_earnings is None:
            raise HTTPException(status_code=400, detail="余额上报必须提供 reported_balance 和 reported_earnings")

        record.reported_balance = data.reported_balance
        record.reported_earnings = data.reported_earnings

        # 计算转入/转出净额 = (新余额 - 旧余额) - (新收益 - 旧收益)
        balance_diff = data.reported_balance - platform.current_balance
        earnings_diff = data.reported_earnings - platform.total_earnings
        record.calculated_transfer = balance_diff - earnings_diff

        # 更新平台数据（total_earnings存当年收益）
        platform.current_balance = data.reported_balance
        platform.total_earnings = data.reported_earnings

    elif data.record_type == "earnings":
        # 仅收益上报模式：填本次收益变化值，自动更新累计收益和余额
        if data.amount is None:
            raise HTTPException(status_code=400, detail="收益上报必须提供 amount（本次收益变化值）")
        new_earnings = platform.total_earnings + data.amount
        record.reported_earnings = new_earnings
        record.reported_balance = platform.current_balance + data.amount
        record.amount = data.amount
        record.calculated_transfer = Decimal('0')
        platform.total_earnings = new_earnings
        platform.current_balance = record.reported_balance

    elif data.record_type == "balance_only":
        # 仅余额上报模式：填余额，收益不变
        if data.reported_balance is None:
            raise HTTPException(status_code=400, detail="余额上报必须提供 reported_balance")
        record.reported_balance = data.reported_balance
        record.reported_earnings = platform.total_earnings
        record.calculated_transfer = data.reported_balance - platform.current_balance
        platform.current_balance = data.reported_balance

    elif data.record_type == "transfer_in":
        if data.amount is None or data.amount <= 0:
            raise HTTPException(status_code=400, detail="转入金额必须大于0")
        record.amount = data.amount
        record.calculated_transfer = data.amount
        platform.current_balance += data.amount

    elif data.record_type == "transfer_out":
        if data.amount is None or data.amount <= 0:
            raise HTTPException(status_code=400, detail="转出金额必须大于0")
        if data.amount > platform.current_balance:
            raise HTTPException(status_code=400, detail="转出金额不能超过当前余额")
        record.amount = data.amount
        record.calculated_transfer = -data.amount
        platform.current_balance -= data.amount

    else:
        raise HTTPException(status_code=400, detail=f"不支持的记录类型: {data.record_type}")

    record.balance_after = platform.current_balance
    record.earnings_after = platform.total_earnings
    platform.updated_at = datetime.now()

    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@router.put("/assets/records/{record_id}", response_model=AssetRecordResponse)
async def update_asset_record(
    record_id: int,
    data: AssetRecordUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """编辑资产变动记录（仅编辑备注和数值，不重新计算平台余额）"""
    record = db.query(AssetRecord).filter(
        AssetRecord.id == record_id,
        AssetRecord.owner_id == current_user.id
    ).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")

    update_data = data.model_dump(exclude_unset=True, exclude_none=True)
    if not update_data:
        raise HTTPException(status_code=400, detail="没有需要更新的字段")

    for key, value in update_data.items():
        setattr(record, key, value)

    db.commit()
    db.refresh(record)
    if record.platform:
        record.platform_name = record.platform.name
    return record


@router.get("/assets/records", response_model=list[AssetRecordResponse])
async def list_asset_records(
    platform_id: Optional[int] = None,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取资产变动记录"""
    query = db.query(AssetRecord).filter(AssetRecord.owner_id == current_user.id)
    if platform_id:
        query = query.filter(AssetRecord.platform_id == platform_id)
    records = query.order_by(AssetRecord.created_at.desc()).limit(limit).all()

    result = []
    for r in records:
        resp = AssetRecordResponse.model_validate(r)
        if r.platform:
            resp.platform_name = r.platform.name
        result.append(resp)
    return result


# ==================== 资产总览 ====================

@router.get("/assets/summary", response_model=AssetSummaryResponse)
async def get_asset_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取资产总览（含各平台详情和记录）
    
    total_earnings 返回当前年份的收益（不是累计）
    yearly_earnings 返回历年收益归档
    """
    platforms = db.query(AssetPlatform).filter(
        AssetPlatform.owner_id == current_user.id,
        AssetPlatform.is_active == True
    ).order_by(AssetPlatform.sort_order, AssetPlatform.id).all()

    total_balance = Decimal('0')
    total_current_year_earnings = Decimal('0')
    all_yearly: dict[str, Decimal] = {}
    current_year = datetime.now().year

    platform_details = []

    for p in platforms:
        if p.is_asset:
            total_balance += (p.current_balance or Decimal('0'))
            total_current_year_earnings += (p.total_earnings or Decimal('0'))

        # 合并历年收益（所有平台都参与）
        yearly = _parse_yearly_earnings(p)
        for yk, yv in yearly.items():
            all_yearly[yk] = all_yearly.get(yk, Decimal('0')) + Decimal(str(yv))

        # 获取所有记录（降序排列，前端展示最近的在上）
        records = db.query(AssetRecord).filter(
            AssetRecord.platform_id == p.id
        ).order_by(AssetRecord.created_at.desc()).all()

        record_responses = []
        for r in records:
            resp = AssetRecordResponse.model_validate(r)
            resp.platform_name = p.name
            record_responses.append(resp)

        # 计算年化收益率（当年收益/当前余额*100%）
        annualized = None
        if p.current_balance and p.current_balance > 0 and p.total_earnings is not None:
            annualized = round(p.total_earnings / p.current_balance * Decimal('100'), 2)

        detail = AssetPlatformDetailResponse(
            id=p.id,
            name=p.name,
            current_balance=p.current_balance or Decimal('0'),
            total_earnings=p.total_earnings or Decimal('0'),
            current_year=p.current_year or current_year,
            yearly_earnings=yearly,
            sort_order=p.sort_order,
            is_active=p.is_active,
            is_asset=p.is_asset,
            created_at=p.created_at,
            updated_at=p.updated_at,
            records=record_responses,
            annualized_return=annualized
        )
        platform_details.append(detail)

    # 当前年份的收益也加入历年总和中
    if str(current_year) not in all_yearly:
        all_yearly[str(current_year)] = total_current_year_earnings
    else:
        all_yearly[str(current_year)] = total_current_year_earnings

    return AssetSummaryResponse(
        total_balance=total_balance,
        total_earnings=total_current_year_earnings,
        yearly_earnings=all_yearly,
        current_year=current_year,
        platforms=platform_details
    )


# ==================== 资产趋势 ====================

@router.get("/assets/trend", response_model=AssetTrendResponse)
async def get_asset_trend(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取资产趋势（按天汇总balance记录的总资产和总收益）"""
    platforms = db.query(AssetPlatform).filter(
        AssetPlatform.owner_id == current_user.id,
        AssetPlatform.is_active == True,
        AssetPlatform.is_asset == True
    ).all()

    if not platforms:
        return AssetTrendResponse(points=[], platforms=[])

    platform_ids = [p.id for p in platforms]

    # 获取所有记录（不限类型），按日期排序
    records = db.query(AssetRecord).filter(
        AssetRecord.platform_id.in_(platform_ids),
    ).order_by(AssetRecord.created_at.asc()).all()

    if not records:
        return AssetTrendResponse(points=[], platforms=[])

    # 按天汇总
    from collections import defaultdict
    daily_data: dict[str, dict] = defaultdict(lambda: {'total_balance': Decimal('0'), 'total_earnings': Decimal('0')})
    platform_daily: dict[str, dict] = defaultdict(lambda: {})

    for r in records:
        # 用北京时间（UTC+8）来分组日期
        beijing_time = r.created_at + timedelta(hours=8)
        date_key = beijing_time.strftime('%Y-%m-%d')
        p = next((p for p in platforms if p.id == r.platform_id), None)
        pname = p.name if p else '?'

        # 该平台当天最新的balance/earnings（优先用balance_after，退而用reported_balance）
        balance = r.balance_after or r.reported_balance or Decimal('0')
        earnings = r.earnings_after or r.reported_earnings or Decimal('0')
        platform_daily[date_key][r.platform_id] = {
            'name': pname,
            'balance': balance,
            'earnings': earnings
        }

    # 按时间顺序累计各平台
    all_dates = sorted(platform_daily.keys())

    # 对每个日期，汇总所有平台的当前值
    points = []
    platform_points = []
    running_balances: dict[int, Decimal] = {}
    running_earnings: dict[int, Decimal] = {}
    prev_total_earnings = Decimal('0')

    for date_key in all_dates:
        # 更新每个平台的运行值
        for pid, data in platform_daily[date_key].items():
            running_balances[pid] = data['balance']
            running_earnings[pid] = data['earnings']

        total_balance = sum(running_balances.values())
        total_earnings = sum(running_earnings.values())
        earnings_delta = total_earnings - prev_total_earnings
        prev_total_earnings = total_earnings

        points.append(AssetTrendPoint(
            date=date_key,
            total_balance=total_balance,
            total_earnings=total_earnings,
            earnings_delta=earnings_delta
        ))

        # 各平台趋势
        for pid in running_balances:
            pname = next((p.name for p in platforms if p.id == pid), '?')
            platform_points.append(PlatformTrendPoint(
                date=date_key,
                name=pname,
                balance=running_balances[pid],
                earnings=running_earnings[pid]
            ))

    # 最后追加今天的样本点（取各平台当前最新值，仅对已有记录的平台）
    today_key = (datetime.now() + timedelta(hours=8)).strftime('%Y-%m-%d')
    if today_key not in platform_daily:
        # 仅对前面有过记录的平台更新今天的值
        for pid in list(running_balances.keys()):
            p = next((x for x in platforms if x.id == pid), None)
            if p:
                running_balances[pid] = p.current_balance or Decimal('0')
                running_earnings[pid] = p.total_earnings or Decimal('0')
        total_balance = sum(running_balances.values())
        total_earnings = sum(running_earnings.values())
        earnings_delta = total_earnings - prev_total_earnings if points else total_earnings

        points.append(AssetTrendPoint(
            date=today_key,
            total_balance=total_balance,
            total_earnings=total_earnings,
            earnings_delta=earnings_delta
        ))

        for pid, pname in [(pid, next((x.name for x in platforms if x.id == pid), '?')) for pid in running_balances]:
            platform_points.append(PlatformTrendPoint(
                date=today_key,
                name=pname,
                balance=running_balances[pid],
                earnings=running_earnings[pid]
            ))

    return AssetTrendResponse(
        points=points,
        platforms=platform_points
    )


@router.get("/assets/zhaopingfei-summary", response_model=ZhaopingfeiSummaryResponse)
async def get_zhaopingfei_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取赵平飞转账年度统计"""
    platform = db.query(AssetPlatform).filter(
        AssetPlatform.owner_id == current_user.id,
        AssetPlatform.name == "赵平飞"
    ).first()

    if not platform:
        return ZhaopingfeiSummaryResponse(years=[], total_in=Decimal('0'), total_out=Decimal('0'), total_net=Decimal('0'))

    records = db.query(AssetRecord).filter(
        AssetRecord.platform_id == platform.id
    ).order_by(AssetRecord.created_at.asc()).all()

    from collections import defaultdict
    yearly: dict[str, dict] = defaultdict(lambda: {'in': Decimal('0'), 'out': Decimal('0')})
    total_in = Decimal('0')
    total_out = Decimal('0')

    for r in records:
        year_key = r.created_at.strftime('%Y')
        if r.record_type == 'transfer_in':
            yearly[year_key]['in'] += (r.amount or Decimal('0'))
            total_in += (r.amount or Decimal('0'))
        elif r.record_type == 'transfer_out':
            yearly[year_key]['out'] += (r.amount or Decimal('0'))
            total_out += (r.amount or Decimal('0'))

    years = []
    for y in sorted(yearly.keys()):
        d = yearly[y]
        years.append(ZhaopingfeiYearSummary(
            year=y,
            transfer_in=d['in'],
            transfer_out=d['out'],
            net=d['in'] - d['out']
        ))

    return ZhaopingfeiSummaryResponse(
        years=years,
        total_in=total_in,
        total_out=total_out,
        total_net=total_in - total_out
    )


@router.get("/assets/export")
async def export_assets(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """导出资产数据为Excel文件"""
    from fastapi.responses import StreamingResponse
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import date

    # 获取所有平台和记录
    platforms = db.query(AssetPlatform).filter(
        AssetPlatform.owner_id == current_user.id,
        AssetPlatform.is_active == True
    ).order_by(AssetPlatform.sort_order, AssetPlatform.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "资产记录"

    # 标题样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E8F4FF", end_color="E8F4FF", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    # 表头
    headers = ["日期", "平台", "类型", "余额变化", "收益变化", "转入/转出", "备注"]
    ws.append(headers)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 获取所有记录
    all_records = []
    for p in platforms:
        records = db.query(AssetRecord).filter(
            AssetRecord.platform_id == p.id
        ).order_by(AssetRecord.created_at.desc()).all()
        for r in records:
            all_records.append((p.name, r))

    # 按日期排序
    all_records.sort(key=lambda x: x[1].created_at, reverse=True)

    # 填充数据
    for p_name, r in all_records:
        row = [
            r.created_at.strftime("%Y-%m-%d %H:%M"),
            p_name,
            r.record_type,
            str(r.reported_balance or ""),
            str(r.reported_earnings or ""),
            str(r.calculated_transfer or ""),
            r.notes or ""
        ]
        ws.append(row)

    # 设置列宽
    ws.column_dimensions["A"].width = 18  # 日期
    ws.column_dimensions["B"].width = 15  # 平台
    ws.column_dimensions["C"].width = 12  # 类型
    ws.column_dimensions["D"].width = 12  # 余额变化
    ws.column_dimensions["E"].width = 12  # 收益变化
    ws.column_dimensions["F"].width = 12  # 转入/转出
    ws.column_dimensions["G"].width = 30  # 备注

    # 添加平台汇总sheet
    ws_summary = wb.create_sheet("平台汇总")
    summary_headers = ["平台", "当前余额", "当年收益", "年化收益率"]
    ws_summary.append(summary_headers)
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for p in platforms:
        annualized = None
        if p.current_balance and p.current_balance > 0 and p.total_earnings is not None:
            annualized = round(p.total_earnings / p.current_balance * 100, 2)

        row = [
            p.name,
            float(p.current_balance or 0),
            float(p.total_earnings or 0),
            f"{annualized}%" if annualized is not None else ""
        ]
        ws_summary.append(row)

    ws_summary.column_dimensions["A"].width = 15
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 15
    ws_summary.column_dimensions["D"].width = 12

    # 添加持仓明细sheet
    ws_items = wb.create_sheet("持仓明细")
    item_headers = ["名称", "编号", "持仓金额", "股基%", "债券%", "现金%", "商品%", "固收%", "其他%", "所属平台"]
    ws_items.append(item_headers)
    for col, header in enumerate(item_headers, 1):
        cell = ws_items.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    items = db.query(AssetItem).filter(
        AssetItem.owner_id == current_user.id
    ).order_by(AssetItem.platform_id, AssetItem.sort_order).all()

    for item in items:
        row = [
            item.name,
            item.code or "",
            float(item.amount or 0),
            float(item.stock_pct or 0),
            float(item.bond_pct or 0),
            float(item.cash_pct or 0),
            float(item.commodity_pct or 0),
            float(item.fixed_income_pct or 0),
            float(item.other_pct or 0),
            item.platform.name if item.platform else ""
        ]
        ws_items.append(row)

    ws_items.column_dimensions["A"].width = 28  # 名称
    ws_items.column_dimensions["B"].width = 12  # 编号
    ws_items.column_dimensions["C"].width = 12  # 持仓金额
    for col_letter in ["D", "E", "F", "G", "H", "I"]:
        ws_items.column_dimensions[col_letter].width = 10  # 占比
    ws_items.column_dimensions["J"].width = 15  # 所属平台

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    filename = f"资产记录_{date.today().strftime('%Y%m%d')}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=utf-8''{encoded_filename}"}
    )


# ==================== 持仓明细（资产项） ====================


@router.get("/assets/items", response_model=list[AssetItemResponse])
async def list_asset_items(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取当前用户的所有持仓项"""
    items = db.query(AssetItem).filter(
        AssetItem.owner_id == current_user.id
    ).order_by(AssetItem.sort_order, AssetItem.id).all()

    result = []
    for item in items:
        resp = AssetItemResponse.model_validate(item)
        if item.platform:
            resp.platform_name = item.platform.name
        result.append(resp)
    return result


@router.post("/assets/items", response_model=AssetItemResponse, status_code=201)
async def create_asset_item(
    data: AssetItemCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """创建持仓项"""
    # 校验比例总和 = 100
    total_pct = data.stock_pct + data.bond_pct + data.cash_pct + data.commodity_pct + data.fixed_income_pct + data.other_pct
    if total_pct != Decimal('100'):
        raise HTTPException(status_code=400, detail=f"各类型占比之和必须为100，当前为{total_pct}")

    item = AssetItem(
        name=data.name,
        code=data.code,
        amount=data.amount,
        stock_pct=data.stock_pct,
        bond_pct=data.bond_pct,
        cash_pct=data.cash_pct,
        commodity_pct=data.commodity_pct,
        fixed_income_pct=data.fixed_income_pct,
        other_pct=data.other_pct,
        platform_id=data.platform_id,
        sort_order=data.sort_order,
        owner_id=current_user.id
    )
    db.add(item)
    db.commit()
    db.refresh(item)

    resp = AssetItemResponse.model_validate(item)
    if item.platform:
        resp.platform_name = item.platform.name
    return resp


@router.put("/assets/items/{item_id}", response_model=AssetItemResponse)
async def update_asset_item(
    item_id: int,
    data: AssetItemUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """更新持仓项"""
    item = db.query(AssetItem).filter(
        AssetItem.id == item_id,
        AssetItem.owner_id == current_user.id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="持仓项不存在")

    update_data = data.model_dump(exclude_unset=True)

    # 如果更新了比例字段，校验总和
    pct_fields = ['stock_pct', 'bond_pct', 'cash_pct', 'commodity_pct', 'fixed_income_pct', 'other_pct']
    has_pct = any(f in update_data for f in pct_fields)
    if has_pct:
        stock = update_data.get('stock_pct', item.stock_pct)
        bond = update_data.get('bond_pct', item.bond_pct)
        cash = update_data.get('cash_pct', item.cash_pct)
        commodity = update_data.get('commodity_pct', item.commodity_pct)
        fixed_income = update_data.get('fixed_income_pct', item.fixed_income_pct)
        other = update_data.get('other_pct', item.other_pct)
        total = stock + bond + cash + commodity + fixed_income + other
        if total != Decimal('100'):
            raise HTTPException(status_code=400, detail=f"各类型占比之和必须为100，当前为{total}")

    for key, value in update_data.items():
        setattr(item, key, value)
    item.updated_at = datetime.now()
    db.commit()
    db.refresh(item)

    resp = AssetItemResponse.model_validate(item)
    if item.platform:
        resp.platform_name = item.platform.name
    return resp


@router.delete("/assets/items/{item_id}")
async def delete_asset_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """删除持仓项"""
    item = db.query(AssetItem).filter(
        AssetItem.id == item_id,
        AssetItem.owner_id == current_user.id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="持仓项不存在")
    db.delete(item)
    db.commit()
    return {"message": "持仓项已删除"}


@router.get("/assets/platform-items", response_model=dict)
async def get_platform_items(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取按平台分组的持仓项，自动计算每个标的占该平台余额的比例"""
    items = db.query(AssetItem).filter(
        AssetItem.owner_id == current_user.id
    ).order_by(AssetItem.platform_id, AssetItem.sort_order, AssetItem.id).all()

    # 获取平台余额
    platforms = db.query(AssetPlatform).filter(
        AssetPlatform.owner_id == current_user.id
    ).all()
    platform_balances = {p.id: p.current_balance or Decimal('0') for p in platforms}

    # 按平台分组
    platform_groups = {}
    for item in items:
        pid = item.platform_id or 0
        if pid not in platform_groups:
            platform_groups[pid] = {
                "platform_id": item.platform.id if item.platform else None,
                "platform_name": item.platform.name if item.platform else "未分配",
                "platform_balance": float(platform_balances.get(pid, Decimal('0'))),
                "items": []
            }

        balance = platform_balances.get(pid, Decimal('0'))
        item_pct = round(float(item.amount) / float(balance) * 100, 2) if balance > 0 else 0

        resp = AssetItemResponse.model_validate(item)
        if item.platform:
            resp.platform_name = item.platform.name
        platform_groups[pid]["items"].append({
            "id": item.id,
            "name": item.name,
            "code": item.code,
            "amount": float(item.amount or 0),
            "stock_pct": float(item.stock_pct or 0),
            "bond_pct": float(item.bond_pct or 0),
            "cash_pct": float(item.cash_pct or 0),
            "commodity_pct": float(item.commodity_pct or 0),
            "fixed_income_pct": float(item.fixed_income_pct or 0),
            "other_pct": float(item.other_pct or 0),
            "platform_name": item.platform.name if item.platform else None,
            "pct_of_platform": item_pct
        })

    return {
        "platforms": [
            {
                "platform_id": g["platform_id"],
                "platform_name": g["platform_name"],
                "platform_balance": g["platform_balance"],
                "items": g["items"]
            }
            for g in sorted(platform_groups.values(), key=lambda x: x["platform_name"] or "")
        ]
    }


@router.get("/assets/portfolio-summary", response_model=PortfolioSummaryResponse)
async def get_portfolio_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取资产组合汇总（按所有持仓项的金额加权计算总比例）"""
    items = db.query(AssetItem).filter(
        AssetItem.owner_id == current_user.id
    ).all()

    total_amount = Decimal('0')
    stock_amount = Decimal('0')
    bond_amount = Decimal('0')
    cash_amount = Decimal('0')
    commodity_amount = Decimal('0')
    fixed_income_amount = Decimal('0')
    other_amount = Decimal('0')

    for item in items:
        amt = item.amount or Decimal('0')
        total_amount += amt
        stock_amount += amt * (item.stock_pct or Decimal('0')) / Decimal('100')
        bond_amount += amt * (item.bond_pct or Decimal('0')) / Decimal('100')
        cash_amount += amt * (item.cash_pct or Decimal('0')) / Decimal('100')
        commodity_amount += amt * (item.commodity_pct or Decimal('0')) / Decimal('100')
        fixed_income_amount += amt * (item.fixed_income_pct or Decimal('0')) / Decimal('100')
        other_amount += amt * (item.other_pct or Decimal('0')) / Decimal('100')

    if total_amount > 0:
        return PortfolioSummaryResponse(
            total_amount=round(total_amount, 2),
            stock_amount=round(stock_amount, 2),
            bond_amount=round(bond_amount, 2),
            cash_amount=round(cash_amount, 2),
            commodity_amount=round(commodity_amount, 2),
            fixed_income_amount=round(fixed_income_amount, 2),
            other_amount=round(other_amount, 2),
            stock_pct=round(stock_amount / total_amount * Decimal('100'), 2),
            bond_pct=round(bond_amount / total_amount * Decimal('100'), 2),
            cash_pct=round(cash_amount / total_amount * Decimal('100'), 2),
            commodity_pct=round(commodity_amount / total_amount * Decimal('100'), 2),
            fixed_income_pct=round(fixed_income_amount / total_amount * Decimal('100'), 2),
            other_pct=round(other_amount / total_amount * Decimal('100'), 2),
        )

    return PortfolioSummaryResponse()


# ==================== 资产快照 ====================


@router.post("/assets/snapshots", response_model=AssetSnapshotResponse, status_code=201)
async def create_asset_snapshot(
    data: AssetSnapshotCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """创建当前持仓快照"""
    from datetime import date as date_type

    # 获取当前持仓数据
    items = db.query(AssetItem).filter(
        AssetItem.owner_id == current_user.id
    ).order_by(AssetItem.platform_id, AssetItem.sort_order).all()

    platforms = db.query(AssetPlatform).filter(
        AssetPlatform.owner_id == current_user.id,
        AssetPlatform.is_asset == True
    ).all()
    platform_balances = {p.id: float(p.current_balance or 0) for p in platforms}

    # 构建持仓明细快照
    snapshot_items = []
    total_from_items = Decimal('0')
    for item in items:
        amt = item.amount or Decimal('0')
        total_from_items += amt
        snapshot_items.append({
            "id": item.id,
            "name": item.name,
            "code": item.code,
            "amount": float(amt),
            "stock_pct": float(item.stock_pct or 0),
            "bond_pct": float(item.bond_pct or 0),
            "cash_pct": float(item.cash_pct or 0),
            "commodity_pct": float(item.commodity_pct or 0),
            "fixed_income_pct": float(item.fixed_income_pct or 0),
            "other_pct": float(item.other_pct or 0),
            "platform_id": item.platform_id,
            "platform_name": item.platform.name if item.platform else None
        })

    snapshot_data = json.dumps(snapshot_items, ensure_ascii=False)
    platform_summary = json.dumps(platform_balances, ensure_ascii=False)

    # 总资产用平台余额之和（更准确）
    total_amount = sum((Decimal(str(v)) for v in platform_balances.values()), Decimal('0'))

    today = date_type.today()
    snapshot = AssetSnapshot(
        owner_id=current_user.id,
        snapshot_date=today,
        snapshot_data=snapshot_data,
        total_amount=round(total_amount, 2),
        platform_summary=platform_summary,
        notes=data.notes
    )
    db.add(snapshot)
    db.commit()
    db.refresh(snapshot)

    return AssetSnapshotResponse.model_validate(snapshot)


@router.get("/assets/snapshots", response_model=list[AssetSnapshotResponse])
async def list_asset_snapshots(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取所有资产快照列表（最新的在前）"""
    snapshots = db.query(AssetSnapshot).filter(
        AssetSnapshot.owner_id == current_user.id
    ).order_by(AssetSnapshot.snapshot_date.desc(), AssetSnapshot.id.desc()).all()

    return [AssetSnapshotResponse.model_validate(s) for s in snapshots]


@router.get("/assets/snapshots/{snapshot_id}", response_model=AssetSnapshotResponse)
async def get_asset_snapshot(
    snapshot_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取单个快照详情"""
    snapshot = db.query(AssetSnapshot).filter(
        AssetSnapshot.id == snapshot_id,
        AssetSnapshot.owner_id == current_user.id
    ).first()
    if not snapshot:
        raise HTTPException(status_code=404, detail="快照不存在")
    return AssetSnapshotResponse.model_validate(snapshot)


# ==================== 固定资产 ====================


@router.get("/assets/fixed-assets", response_model=list[FixedAssetResponse])
async def list_fixed_assets(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取所有固定资产"""
    assets = db.query(FixedAsset).filter(
        FixedAsset.owner_id == current_user.id
    ).order_by(FixedAsset.sort_order, FixedAsset.id).all()
    return [FixedAssetResponse.model_validate(a) for a in assets]


@router.post("/assets/fixed-assets", response_model=FixedAssetResponse, status_code=201)
async def create_fixed_asset(
    data: FixedAssetCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """创建固定资产"""
    asset = FixedAsset(
        owner_id=current_user.id,
        name=data.name,
        category=data.category,
        estimated_value=data.estimated_value,
        role=data.role,
        monthly_rent=data.monthly_rent,
        notes=data.notes,
        sort_order=data.sort_order
    )
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return FixedAssetResponse.model_validate(asset)


@router.put("/assets/fixed-assets/{asset_id}", response_model=FixedAssetResponse)
async def update_fixed_asset(
    asset_id: int,
    data: FixedAssetUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """更新固定资产"""
    asset = db.query(FixedAsset).filter(
        FixedAsset.id == asset_id,
        FixedAsset.owner_id == current_user.id
    ).first()
    if not asset:
        raise HTTPException(status_code=404, detail="固定资产不存在")

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(asset, key, value)
    db.commit()
    db.refresh(asset)
    return FixedAssetResponse.model_validate(asset)


@router.delete("/assets/fixed-assets/{asset_id}")
async def delete_fixed_asset(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """删除固定资产"""
    asset = db.query(FixedAsset).filter(
        FixedAsset.id == asset_id,
        FixedAsset.owner_id == current_user.id
    ).first()
    if not asset:
        raise HTTPException(status_code=404, detail="固定资产不存在")
    db.delete(asset)
    db.commit()
    return {"message": "固定资产已删除"}
