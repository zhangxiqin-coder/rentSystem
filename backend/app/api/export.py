# Data export endpoints
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime
import csv
import io
from io import StringIO

from app.database import get_db
from app.models import Room, Payment, UtilityReading, Tenant, LeaseRecord
from app.core.deps import get_current_user
from app.models import User

router = APIRouter(prefix="/export", tags=["export"])


@router.get("/utility-readings")
async def export_utility_readings(
    start_date: str = None,
    end_date: str = None,
    room_id: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Export utility readings to CSV format.
    Query params:
    - start_date: Start date filter (YYYY-MM-DD)
    - end_date: End date filter (YYYY-MM-DD)
    - room_id: Filter by room ID
    """
    query = db.query(UtilityReading)

    if start_date:
        query = query.filter(UtilityReading.reading_date >= start_date)
    if end_date:
        query = query.filter(UtilityReading.reading_date <= end_date)
    if room_id:
        query = query.filter(UtilityReading.room_id == room_id)

    readings = query.order_by(UtilityReading.reading_date.desc(), UtilityReading.room_id).all()

    # Create CSV
    output = StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        '房间号', '抄表日期', '类型', '上次读数', '本次读数', '用量',
        '单价', '金额', '备注', '创建时间'
    ])

    # Data rows
    for reading in readings:
        room = db.query(Room).filter(Room.id == reading.room_id).first()
        room_number = room.room_number if room else f"房间{reading.room_id}"

        writer.writerow([
            room_number,
            reading.reading_date,
            '水' if reading.utility_type == 'water' else '电',
            reading.previous_reading,
            reading.reading,
            reading.usage,
            reading.rate,
            f"{reading.amount or 0:.2f}",
            reading.notes or '',
            reading.created_at.strftime('%Y-%m-%d %H:%M:%S') if reading.created_at else ''
        ])

    # Generate filename
    filename = f"utility_readings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    output.seek(0)
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/payments")
async def export_payments(
    start_date: str = None,
    end_date: str = None,
    room_id: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Export payment records to CSV format.
    Query params:
    - start_date: Start date filter (YYYY-MM-DD)
    - end_date: End date filter (YYYY-MM-DD)
    - room_id: Filter by room ID
    """
    query = db.query(Payment)

    if start_date:
        query = query.filter(Payment.payment_date >= start_date)
    if end_date:
        query = query.filter(Payment.payment_date <= end_date)
    if room_id:
        query = query.filter(Payment.room_id == room_id)

    payments = query.order_by(Payment.payment_date.desc()).all()

    # Create CSV
    output = StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        '房间号', '收款日期', '房租', '水费', '电费', '总金额',
        '收款方式', '备注', '创建时间'
    ])

    # Data rows
    for payment in payments:
        room = db.query(Room).filter(Room.id == payment.room_id).first()
        room_number = room.room_number if room else f"房间{payment.room_id}"

        # Extract utility charges
        water_amount = 0
        electricity_amount = 0

        if payment.utility_charges:
            for charge in payment.utility_charges:
                if charge.utility_type == 'water':
                    water_amount = charge.amount or 0
                elif charge.utility_type == 'electricity':
                    electricity_amount = charge.amount or 0

        total = (payment.amount or 0) + water_amount + electricity_amount

        writer.writerow([
            room_number,
            payment.payment_date,
            f"{payment.amount or 0:.2f}",
            f"{water_amount:.2f}",
            f"{electricity_amount:.2f}",
            f"{total:.2f}",
            payment.payment_method,
            payment.notes or '',
            payment.created_at.strftime('%Y-%m-%d %H:%M:%S') if payment.created_at else ''
        ])

    # Generate filename
    filename = f"payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    output.seek(0)
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/rooms")
async def export_rooms(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Export room information to CSV format.
    """
    rooms = db.query(Room).order_by(Room.room_number).all()

    # Create CSV
    output = StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        '房间号', '楼层', '月租金', '付款周期（月）', '租约开始', '租约结束',
        '上次付款日期', '水费率', '电费率', '状态', '创建时间'
    ])

    # Data rows
    for room in rooms:
        status = '已租' if room.is_occupied else '空置'

        writer.writerow([
            room.room_number,
            room.floor,
            f"{room.monthly_rent:.2f}",
            room.payment_cycle,
            room.lease_start,
            room.lease_end,
            room.last_payment_date or '',
            f"{room.water_rate:.2f}",
            f"{room.electricity_rate:.2f}",
            status,
            room.created_at.strftime('%Y-%m-%d %H:%M:%S') if room.created_at else ''
        ])

    # Generate filename
    filename = f"rooms_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    output.seek(0)
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/system-overview")
async def export_system_overview(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    导出系统全量数据（Excel格式），按系列→房间号排序。
    包含房间信息、租客个人信息（身份证等）、上次水电读数。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # 获取当前用户的所有房间，按系列→房间号排序
    rooms = db.query(Room).filter(
        Room.owner_id == current_user.id
    ).order_by(Room.series, Room.room_number).all()

    # 预加载：每个房间的最新水电读数
    room_ids = [r.id for r in rooms]

    latest_water = {}
    latest_electric = {}
    if room_ids:
        # 使用子查询找每个房间最新的水/电读数
        for ut_type in ('water', 'electricity'):
            sub = db.query(
                UtilityReading.room_id,
                func.max(UtilityReading.reading_date).label('max_date')
            ).filter(
                UtilityReading.room_id.in_(room_ids),
                UtilityReading.utility_type == ut_type
            ).group_by(UtilityReading.room_id).subquery()

            readings = db.query(UtilityReading).join(
                sub,
                (UtilityReading.room_id == sub.c.room_id) &
                (UtilityReading.reading_date == sub.c.max_date) &
                (UtilityReading.utility_type == ut_type)
            ).all()

            for r in readings:
                if ut_type == 'water':
                    latest_water[r.room_id] = r
                else:
                    latest_electric[r.room_id] = r

    # 预加载租客信息（通过 room.tenant_id 关联）
    tenant_ids = [r.tenant_id for r in rooms if r.tenant_id]
    tenants_map = {}
    if tenant_ids:
        tenants = db.query(Tenant).filter(Tenant.id.in_(tenant_ids)).all()
        tenants_map = {t.id: t for t in tenants}

    # 创建Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "房间数据导出"

    # 样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 表头
    headers = [
        '房间号', '楼栋', '系列', '租金', '水费率', '电费率', '付款周期',
        '租客姓名', '租客电话', '身份证号', '紧急联系人', '紧急联系电话',
        '租约开始', '租约结束',
        '上次水表读数', '上次电表读数', '上次抄表日期'
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 数据行
    for room in rooms:
        tenant = tenants_map.get(room.tenant_id) if room.tenant_id else None
        water_r = latest_water.get(room.id)
        electric_r = latest_electric.get(room.id)

        # 上次抄表日期取水电中较新的
        water_date = water_r.reading_date if water_r else None
        electric_date = electric_r.reading_date if electric_r else None
        last_reading_date = max(water_date, electric_date) if (water_date or electric_date) else None

        ws.append([
            room.room_number or '',
            room.building or '',
            room.series or '',
            float(room.monthly_rent) if room.monthly_rent else 0,
            float(room.water_rate) if room.water_rate else 0,
            float(room.electricity_rate) if room.electricity_rate else 0,
            room.payment_cycle or 1,
            tenant.name.strip() if tenant and tenant.name else (room.tenant_name or ''),
            tenant.phone if tenant and tenant.phone else (room.tenant_phone or ''),
            tenant.id_card if tenant and tenant.id_card else '',
            tenant.emergency_contact if tenant and tenant.emergency_contact else '',
            tenant.emergency_phone if tenant and tenant.emergency_phone else '',
            room.lease_start or '',
            room.lease_end or '',
            float(water_r.reading) if water_r else '',
            float(electric_r.reading) if electric_r else '',
            last_reading_date or '',
        ])

    # 设置列宽
    col_widths = [10, 8, 10, 10, 8, 8, 10, 10, 14, 22, 10, 14, 12, 12, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    # 添加自动筛选
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(headers)).column_letter}{len(rooms) + 1}"

    # 输出
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    filename = f"系统数据导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )
